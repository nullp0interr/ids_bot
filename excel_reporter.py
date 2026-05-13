import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_ansible_report(test_results, test_list, filename="report.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ошибки авторизации"

    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))

    # Заголовки
    ws.append(["Имя забикса", "Ожидаемый IP", "Статус"])

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Данные (добавляем ТОЛЬКО ТЕ, ГДЕ СТАТУС NO)
    failures_found = False
    for zabbix_name, status in test_results.items():
        if status == "NO":
            failures_found = True
            expected_ip = test_list.get(zabbix_name, "забикс пуст")
            ws.append([zabbix_name, expected_ip, status])
            
            for cell in ws[ws.max_row]:
                cell.fill = red_fill
                cell.font = red_font
                cell.border = thin_border
                cell.alignment = left_align
            ws.cell(row=ws.max_row, column=3).alignment = center_align

    # Если ошибок нет, добавляем радостную строчку
    if not failures_found:
        ws.append(["прошли проверку", "", ""])
        ws.merge_cells("A2:C2")
        ws.cell(row=2, column=1).alignment = center_align
        ws.cell(row=2, column=1).font = Font(bold=True, color="2E7D32")

    # Авто-ширина колонок
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

    wb.save(filename)
    return filename